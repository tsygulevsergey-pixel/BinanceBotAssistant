import openpyxl

def extract_metrics():
    """Извлечение ключевых метрик из TradingView отчетов"""
    
    files = {
        'BTCUSDT': '../attached_assets/Smart_Money_Liquidity_Hunt_BINANCE_BTCUSDT.P_2025-10-13_846cf_1760382115223.xlsx',
        'ETHUSDT': '../attached_assets/Smart_Money_Liquidity_Hunt_BINANCE_ETHUSDT.P_2025-10-13_98d79_1760382152153.xlsx',
        'SOLUSDT': '../attached_assets/Smart_Money_Liquidity_Hunt_BINANCE_SOLUSDT.P_2025-10-13_672c5_1760382187508.xlsx'
    }
    
    results = {}
    
    for symbol, filepath in files.items():
        wb = openpyxl.load_workbook(filepath)
        metrics = {}
        
        # Читаем лист "Динамика"
        dynamics = wb['Динамика']
        for row in dynamics.iter_rows(values_only=True):
            if row[0] == 'Чистая прибыль':
                metrics['net_profit'] = row[1]
                metrics['net_profit_pct'] = row[2]
        
        # Читаем лист "Анализ сделок"
        analysis = wb['Анализ сделок']
        for row in analysis.iter_rows(values_only=True):
            if row[0] == 'Всего сделок':
                metrics['total_trades'] = row[1]
            elif row[0] == 'Прибыльные сделки':
                metrics['winning_trades'] = row[1]
            elif row[0] == 'Убыточные сделки':
                metrics['losing_trades'] = row[1]
            elif row[0] == 'Процент прибыльных':
                metrics['win_rate'] = row[2]
            elif row[0] == 'Средние ПР/УБ':
                metrics['avg_trade'] = row[1]
                metrics['avg_trade_pct'] = row[2]
        
        # Читаем лист "Коэффициенты риска"
        risk = wb['Коэффициенты риска эффективн...']
        for row in risk.iter_rows(values_only=True):
            if row[0] == 'Фактор прибыли':
                metrics['profit_factor'] = row[1]
        
        results[symbol] = metrics
        wb.close()
    
    # Вывод отчета
    print("\n" + "="*80)
    print("📊 ПОЛНЫЙ АНАЛИЗ: Smart Money Liquidity Hunt Strategy")
    print("="*80)
    
    for symbol, m in results.items():
        print(f"\n{'─'*80}")
        print(f"🎯 {symbol}")
        print(f"{'─'*80}")
        print(f"Чистая прибыль:      {m.get('net_profit', 'N/A'):>10} USDT ({m.get('net_profit_pct', 'N/A'):>6}%)")
        print(f"Всего сделок:        {m.get('total_trades', 'N/A'):>10}")
        print(f"Прибыльных сделок:   {m.get('winning_trades', 'N/A'):>10}")
        print(f"Убыточных сделок:    {m.get('losing_trades', 'N/A'):>10}")
        print(f"Win Rate:            {m.get('win_rate', 'N/A'):>10}%")
        print(f"Profit Factor:       {m.get('profit_factor', 'N/A'):>10}")
        print(f"Средняя сделка:      {m.get('avg_trade', 'N/A'):>10} USDT ({m.get('avg_trade_pct', 'N/A'):>6}%)")
    
    # Сравнительный анализ
    print(f"\n{'='*80}")
    print("📈 СРАВНИТЕЛЬНЫЙ АНАЛИЗ")
    print(f"{'='*80}\n")
    
    print(f"{'Метрика':<25} {'BTCUSDT':>15} {'ETHUSDT':>15} {'SOLUSDT':>15}")
    print(f"{'-'*80}")
    print(f"{'Win Rate %':<25} {results['BTCUSDT'].get('win_rate', 'N/A'):>15} {results['ETHUSDT'].get('win_rate', 'N/A'):>15} {results['SOLUSDT'].get('win_rate', 'N/A'):>15}")
    print(f"{'Profit Factor':<25} {results['BTCUSDT'].get('profit_factor', 'N/A'):>15} {results['ETHUSDT'].get('profit_factor', 'N/A'):>15} {results['SOLUSDT'].get('profit_factor', 'N/A'):>15}")
    print(f"{'Net Profit %':<25} {results['BTCUSDT'].get('net_profit_pct', 'N/A'):>15} {results['ETHUSDT'].get('net_profit_pct', 'N/A'):>15} {results['SOLUSDT'].get('net_profit_pct', 'N/A'):>15}")
    print(f"{'Total Trades':<25} {results['BTCUSDT'].get('total_trades', 'N/A'):>15} {results['ETHUSDT'].get('total_trades', 'N/A'):>15} {results['SOLUSDT'].get('total_trades', 'N/A'):>15}")
    print(f"{'Avg Trade %':<25} {results['BTCUSDT'].get('avg_trade_pct', 'N/A'):>15} {results['ETHUSDT'].get('avg_trade_pct', 'N/A'):>15} {results['SOLUSDT'].get('avg_trade_pct', 'N/A'):>15}")
    
    # Выводы
    print(f"\n{'='*80}")
    print("💡 КЛЮЧЕВЫЕ ВЫВОДЫ")
    print(f"{'='*80}\n")
    
    win_rates = [results[s].get('win_rate', 0) for s in results]
    profit_factors = [results[s].get('profit_factor', 0) for s in results]
    
    avg_win_rate = sum(win_rates) / len(win_rates)
    avg_profit_factor = sum(profit_factors) / len(profit_factors)
    
    print(f"✅ Средний Win Rate:      {avg_win_rate:.1f}%")
    print(f"✅ Средний Profit Factor: {avg_profit_factor:.2f}")
    
    # Универсальность
    win_rate_std = (max(win_rates) - min(win_rates))
    if win_rate_std < 15:
        print(f"\n🎯 УНИВЕРСАЛЬНОСТЬ: ОТЛИЧНО (разброс Win Rate: {win_rate_std:.1f}%)")
        print("   Стратегия работает стабильно на всех парах!")
    elif win_rate_std < 25:
        print(f"\n🟡 УНИВЕРСАЛЬНОСТЬ: СРЕДНЕ (разброс Win Rate: {win_rate_std:.1f}%)")
        print("   Есть различия между парами, требуется оптимизация")
    else:
        print(f"\n❌ УНИВЕРСАЛЬНОСТЬ: ПЛОХО (разброс Win Rate: {win_rate_std:.1f}%)")
        print("   Стратегия overfitted на одну пару")
    
    # Общая оценка
    print(f"\n{'='*80}")
    print("🏆 ОБЩАЯ ОЦЕНКА")
    print(f"{'='*80}\n")
    
    if avg_win_rate >= 50 and avg_profit_factor >= 1.3:
        print("✅ РЕЗУЛЬТАТ: ОТЛИЧНО!")
        print("   Стратегия прибыльна и универсальна")
        print("   Рекомендация: Использовать в боте")
    elif avg_win_rate >= 45 and avg_profit_factor >= 1.1:
        print("🟡 РЕЗУЛЬТАТ: ХОРОШО")
        print("   Стратегия прибыльна, но есть потенциал для улучшения")
        print("   Рекомендация: Оптимизировать параметры")
    else:
        print("❌ РЕЗУЛЬТАТ: ТРЕБУЕТ ДОРАБОТКИ")
        print("   Стратегия показывает слабые результаты")
        print("   Рекомендация: Пересмотреть концепцию")
    
    print(f"\n{'='*80}\n")
    
    return results

if __name__ == "__main__":
    extract_metrics()
