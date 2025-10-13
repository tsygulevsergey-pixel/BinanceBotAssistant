import openpyxl
import pandas as pd
from pathlib import Path

def analyze_strategy_results():
    """Анализ результатов тестирования стратегии Smart Money Liquidity Hunt"""
    
    files = {
        'BTCUSDT': '../attached_assets/Smart_Money_Liquidity_Hunt_BINANCE_BTCUSDT.P_2025-10-13_846cf_1760382115223.xlsx',
        'ETHUSDT': '../attached_assets/Smart_Money_Liquidity_Hunt_BINANCE_ETHUSDT.P_2025-10-13_98d79_1760382152153.xlsx',
        'SOLUSDT': '../attached_assets/Smart_Money_Liquidity_Hunt_BINANCE_SOLUSDT.P_2025-10-13_672c5_1760382187508.xlsx'
    }
    
    results = {}
    
    for symbol, filepath in files.items():
        try:
            # Читаем Excel файл
            wb = openpyxl.load_workbook(filepath)
            
            # Выводим все листы
            print(f"\n{'='*60}")
            print(f"АНАЛИЗ: {symbol}")
            print(f"{'='*60}")
            print(f"Доступные листы: {wb.sheetnames}")
            
            # Читаем основной лист (обычно первый)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                print(f"\n--- Лист: {sheet_name} ---")
                
                # Выводим первые 30 строк
                for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    if i > 30:
                        break
                    print(row)
                
            wb.close()
            
        except Exception as e:
            print(f"Ошибка при чтении {symbol}: {e}")
    
    print("\n" + "="*60)
    print("АНАЛИЗ ЗАВЕРШЕН")
    print("="*60)

if __name__ == "__main__":
    analyze_strategy_results()
